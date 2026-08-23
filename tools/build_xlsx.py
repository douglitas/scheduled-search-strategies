#!/usr/bin/env python3
"""Genera docs/research.xlsx desde data/*.tsv. Una hoja por pestaña, para
enviar a terceros (un mentor, la directora de tesis, un careers advisor)."""
import csv, os, datetime, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, 'data')
OUT = os.path.join(ROOT, 'docs', 'research.xlsx')
ORDER = ['action_now', 'postdocs', 'jobs', 'fellowships', 'groups',
         'events', 'training', 'watchlist_closed',
         'subscriptions', 'sources', 'inbox_triage', 'changelog',
         'readme']
HEAD = PatternFill('solid', fgColor='2E2C28')

wb = Workbook()
wb.remove(wb.active)
today = datetime.date.today()

for tab in ORDER:
    p = os.path.join(DATA, tab + '.tsv')
    if not os.path.exists(p):
        continue
    rows = list(csv.reader(open(p, newline='', encoding='utf-8'), delimiter='\t'))
    if not rows:
        continue
    ws = wb.create_sheet(tab[:31])
    for r in rows:
        ws.append(r)
    for c in range(1, len(rows[0]) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = HEAD
        cell.alignment = Alignment(vertical='top')
        width = max((len(str(r[c-1])) for r in rows[:200] if len(r) >= c), default=10)
        ws.column_dimensions[get_column_letter(c)].width = min(max(width + 2, 10), 52)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(rows[0]))}{len(rows)}'
    # Days_Left recalculado, no heredado: la columna quedaria obsoleta entre pasadas
    hdr = rows[0]
    if 'Deadline' in hdr and 'Days_Left' in hdr:
        di, li = hdr.index('Deadline') + 1, hdr.index('Days_Left') + 1
        for i in range(2, len(rows) + 1):
            v = str(ws.cell(i, di).value or '')
            if re.match(r'^\d{4}-\d{2}-\d{2}$', v):
                y, m, d = map(int, v.split('-'))
                ws.cell(i, li).value = (datetime.date(y, m, d) - today).days
            else:
                ws.cell(i, li).value = None

wb.save(OUT)
print(f'{OUT}  {os.path.getsize(OUT)} bytes · {len(wb.sheetnames)} hojas: {", ".join(wb.sheetnames)}')
